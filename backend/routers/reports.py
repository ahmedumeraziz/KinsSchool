from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from utils.auth import get_current_user
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

router = APIRouter()

MONTHLY_DATA = [
    {"month": "February",  "collected": 68000, "target": 80000, "pending": 12000},
    {"month": "March",     "collected": 72000, "target": 80000, "pending": 8000},
    {"month": "April",     "collected": 65000, "target": 80000, "pending": 15000},
    {"month": "May",       "collected": 78000, "target": 80000, "pending": 2000},
    {"month": "June",      "collected": 55000, "target": 80000, "pending": 25000},
    {"month": "July",      "collected": 74000, "target": 80000, "pending": 6000},
    {"month": "August",    "collected": 79000, "target": 80000, "pending": 1000},
    {"month": "September", "collected": 82000, "target": 80000, "pending": 0},
    {"month": "October",   "collected": 76000, "target": 80000, "pending": 4000},
    {"month": "November",  "collected": 85000, "target": 80000, "pending": 0},
    {"month": "December",  "collected": 71000, "target": 80000, "pending": 9000},
    {"month": "January",   "collected": 69000, "target": 80000, "pending": 11000},
]

@router.get("/monthly")
async def monthly_report(user=Depends(get_current_user)):
    total_collected = sum(m["collected"] for m in MONTHLY_DATA)
    total_target    = sum(m["target"]    for m in MONTHLY_DATA)
    total_pending   = sum(m["pending"]   for m in MONTHLY_DATA)
    return {
        "months": MONTHLY_DATA,
        "summary": {
            "totalCollected": total_collected,
            "totalTarget": total_target,
            "totalPending": total_pending,
            "recoveryPct": round((total_collected / total_target) * 100, 1) if total_target else 0,
        }
    }

@router.get("/monthly/excel")
async def monthly_excel(user=Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Fee Report"

    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "KINS SCHOOL — Monthly Fee Collection Report"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1B4FD8")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, size=10, color="64748B")

    headers = ["Month", "Target (Rs)", "Collected (Rs)", "Pending (Rs)", "Recovery %"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F2878")
        cell.alignment = Alignment(horizontal="center")

    blue_fill  = PatternFill("solid", fgColor="EFF6FF")
    green_fill = PatternFill("solid", fgColor="ECFDF5")
    red_fill   = PatternFill("solid", fgColor="FEF2F2")

    for row_idx, m in enumerate(MONTHLY_DATA, 5):
        pct = round((m["collected"] / m["target"]) * 100, 1) if m["target"] else 0
        data = [m["month"], m["target"], m["collected"], m["pending"], f"{pct}%"]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")
            if col >= 2:
                cell.number_format = "#,##0"
            if row_idx % 2 == 0:
                cell.fill = blue_fill

    # Totals
    tot_row = len(MONTHLY_DATA) + 5
    totals = ["TOTAL",
              sum(m["target"] for m in MONTHLY_DATA),
              sum(m["collected"] for m in MONTHLY_DATA),
              sum(m["pending"] for m in MONTHLY_DATA),
              f"{round(sum(m['collected'] for m in MONTHLY_DATA)/sum(m['target'] for m in MONTHLY_DATA)*100,1)}%"]
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=tot_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    col_widths = [16, 16, 16, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Monthly-Fee-Report.xlsx"})

@router.get("/monthly/pdf")
async def monthly_pdf(user=Depends(get_current_user)):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("<b>KINS SCHOOL — Monthly Fee Collection Report</b>", styles["Title"]))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Ratta Rd, Kins St, Gujranwala", styles["Normal"]))
    story.append(Spacer(1, 14))

    rows = [["Month", "Target (Rs)", "Collected (Rs)", "Pending (Rs)", "Recovery %"]]
    for m in MONTHLY_DATA:
        pct = round((m["collected"] / m["target"]) * 100, 1) if m["target"] else 0
        rows.append([m["month"], f"{m['target']:,}", f"{m['collected']:,}", f"{m['pending']:,}", f"{pct}%"])

    total_c = sum(m["collected"] for m in MONTHLY_DATA)
    total_t = sum(m["target"]    for m in MONTHLY_DATA)
    total_p = sum(m["pending"]   for m in MONTHLY_DATA)
    rows.append(["TOTAL", f"{total_t:,}", f"{total_c:,}", f"{total_p:,}", f"{round(total_c/total_t*100,1)}%"])

    t = Table(rows, colWidths=[120, 100, 120, 100, 90])
    t.setStyle(TableStyle([
        ("BACKGROUND",      (0, 0), (-1, 0), colors.HexColor("#1B4FD8")),
        ("TEXTCOLOR",       (0, 0), (-1, 0), colors.white),
        ("FONTNAME",        (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",        (0, 0), (-1, -1), 9),
        ("GRID",            (0, 0), (-1, -1), 0.4, colors.lightgrey),
        ("ROWBACKGROUNDS",  (0, 1), (-1, -2), [colors.white, colors.HexColor("#F0F7FF")]),
        ("BACKGROUND",      (0, -1), (-1, -1), colors.HexColor("#ECFDF5")),
        ("FONTNAME",        (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN",           (1, 0), (-1, -1), "CENTER"),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=Monthly-Report.pdf"})

@router.get("/defaulters/excel")
async def defaulters_excel(user=Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defaulters"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"KINS SCHOOL — Defaulter List — {datetime.now().strftime('%B %Y')}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="DC2626")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["#", "Student Name", "Class", "Father Cell", "Unpaid Months", "Amount Due", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="991B1B")
        cell.alignment = Alignment(horizontal="center")

    demo_defaulters = [
        (1, "Ahmed Raza",   "Class 5",  "0300-1234567", "March, April",           3000,  "Strict"),
        (2, "Sara Malik",   "Class 3",  "0311-2345678", "Feb, March, April",      5950,  "Strict"),
        (3, "Bilal Hassan", "Class 6",  "0300-1234567", "Feb, March, April",      5000,  "Strict"),
        (4, "Fatima Khan",  "Class 8",  "0333-4567890", "March, April",           5500,  "Grace"),
        (5, "Zainab Iqbal", "Class 4",  "0344-5678901", "April",                  750,   "First Time"),
    ]

    red_fill = PatternFill("solid", fgColor="FEF2F2")
    for i, row in enumerate(demo_defaulters, 4):
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")
            if i % 2 == 0:
                cell.fill = red_fill

    for i, w in enumerate([5, 20, 12, 16, 24, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Defaulters-List.xlsx"})
